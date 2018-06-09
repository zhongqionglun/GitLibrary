# -*-coding:UTF-8-*-

from openpyxl import Workbook, load_workbook, cell
import datetime

# Create workbook

# Change background color
from openpyxl.styles import PatternFill, Color

wb = Workbook()
# grab active worksheet
ws = wb.active
ws.sheet_properties.tabColor = "ff0000"
ws['A1'].fill = PatternFill(patternType='solid', fill_type='solid', fgColor=Color('ff0000'))
# Data can be assigned directly cells
ws['A1'] = 10
ws['B1'] = 20
ws['C1'] = 30
# Rows can also be
for row in range(1, 40):
    ws.append(range(600))
# Python types will automatically be converted
ws['A2'] = datetime.datetime.now()
# save the file
ws1 = wb.create_sheet()
ws1.title = "secondsheet"
ws1.cell(row=1, column=1, value="成功")
d = ws1.cell(row=2, column=1, value="失败")
# for sheet in wb:
# 	print(wb.sheetnames)
wb.save("test.xlsx")

Workbook1 = load_workbook('test.xlsx')
print(Workbook1.sheetnames)
worksheet2 = Workbook1.sheetnames[1]
print(worksheet2)
sheet1 = Workbook1.get_sheet_by_name(worksheet2)
print(sheet1.cell(row=1, column=1).value)
print(sheet1.cell(row=2, column=1).value)
wb.save("test.xlsx")
